#!/usr/bin/env python3
"""
Extract process data from CostTable_Processes_2025.xls using openpyxl
"""

import openpyxl
import json

print("Reading Process Cost Table with openpyxl...")

try:
    wb = openpyxl.load_workbook('/Users/brajeshkumar/Desktop/SUPRA REPORt/CostTable_Processes_2025.xls')
    ws = wb.active
    
    print(f"Sheet name: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Print first 15 rows to see structure
    print("\nFirst 15 rows:")
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        print(row)
        
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
