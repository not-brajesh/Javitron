#!/usr/bin/env python3
"""
Export SUPRA SAEINDIA Cost Report Workbook to PDF
"""

import os
import subprocess
from openpyxl import load_workbook

def export_excel_to_pdf_macos(excel_file, pdf_file):
    """Export Excel to PDF on macOS using osascript"""
    
    applescript = f'''
    tell application "Microsoft Excel"
        open POSIX file "{os.path.abspath(excel_file)}"
        set theWorkbook to active workbook
        
        tell theWorkbook
            save as filename POSIX file "{os.path.abspath(pdf_file)}" file format PDF file format
            close saving no
        end tell
    end tell
    '''
    
    try:
        subprocess.run(['osascript', '-e', applescript], check=True)
        print(f"Successfully exported to {pdf_file}")
        return True
    except subprocess.CalledProcessError as e:
        print(f"Error exporting to PDF: {e}")
        print("Note: This requires Microsoft Excel for Mac to be installed")
        return False

def verify_workbook_structure(excel_file):
    """Verify the workbook structure"""
    print("Verifying workbook structure...")
    
    wb = load_workbook(excel_file)
    
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print(f"  Sheet names: {wb.sheetnames}")
    
    expected_sheets = [
        'Cover Page',
        'Table of Contents',
        'Cost Summary',
        'BOM',
        'Penalty Calculations'
    ]
    
    for code in ['BR', 'EN', 'FR', 'EL', 'MS', 'ST', 'SU', 'WT']:
        expected_sheets.append(f'{code} - Brakes' if code == 'BR' else 
                               f'{code} - Engine & Drivetrain' if code == 'EN' else
                               f'{code} - Frame & Body' if code == 'FR' else
                               f'{code} - Electrical' if code == 'EL' else
                               f'{code} - Misc' if code == 'MS' else
                               f'{code} - Steering' if code == 'ST' else
                               f'{code} - Suspension' if code == 'SU' else
                               f'{code} - Wheels')
    
    missing_sheets = set(expected_sheets) - set(wb.sheetnames)
    if missing_sheets:
        print(f"  WARNING: Missing sheets: {missing_sheets}")
    else:
        print("  All expected sheets present ✓")
    
    # Verify team identity in Cover Page
    cover = wb['Cover Page']
    university = cover['E5'].value
    team = cover['E7'].value
    
    print(f"  University: {university}")
    print(f"  Team Name: {team}")
    
    if university == "Ajeenkya DY Patil University" and team == "Javitron":
        print("  Team identity correctly populated ✓")
    else:
        print("  WARNING: Team identity may not be correctly set")
    
    wb.close()
    
    return True

def create_user_guide():
    """Create a user guide for the workbook"""
    guide_content = """# SUPRA SAEINDIA Cost Report Workbook - User Guide

## Team: Javitron
## University: Ajeenkya DY Patil University
## Season: 2025-26

### Workbook Structure

The workbook contains the following sheets:

1. **Cover Page** - Team information and report details
2. **Table of Contents** - Navigation guide
3. **Cost Summary** - Overall cost breakdown by commodity
4. **BOM** - Bill of Materials with all parts
5. **Penalty Calculations** - Method A and Method B penalty tracking
6. **Commodity Sheets** (8 sheets):
   - BR - Brakes
   - EN - Engine & Drivetrain
   - FR - Frame & Body
   - EL - Electrical
   - MS - Misc
   - ST - Steering
   - SU - Suspension
   - WT - Wheels

### How to Use

#### 1. Fill in Car Number
- Go to Cover Page sheet
- Enter your car number in cell E9

#### 2. Complete BOM
- Go to BOM sheet
- Fill in all parts with:
  - Part Number (format: XX-YY-001, e.g., FR-FM-001)
  - Part Name
  - Assembly
  - Commodity (BR, EN, FR, EL, MS, ST, SU, WT)
  - Make/Buy (Make or Buy)
  - Quantity
  - Unit Cost (₹)

#### 3. Complete Commodity Sheets
For each commodity sheet:
- Fill in Assembly name
- Add Parts in the Parts section
- Add Materials in the Materials section
- Add Processes in the Processes section
- Add Fasteners in the Fasteners section
- Add Tooling in the Tooling section

#### 4. Part Numbering Format
Follow Appendix C-2: [Commodity Code] - [Assembly Code] - [Part Number]
Example: SU-AR-001 (Suspension - A-Arm - 001)

#### 5. Cost Calculations
All cost calculations are automatic formulas:
- Parts: Sub Total = Part Cost × Quantity
- Materials: Sub Total = Unit Cost × Quantity
- Processes: Sub Total = Unit Cost × Quantity × Multiplier
- Fasteners: Sub Total = Unit Cost × Quantity
- Tooling: Sub Total = Unit Cost × Quantity × PVF × Fraction Included

#### 6. Validation
- Red highlight = Error (missing required field)
- Yellow highlight = Warning (incomplete data)
- Fill all red cells before submission

#### 7. Made Parts
For parts made from raw material:
- Specify raw material type, shape, dimensions
- Provide gross weight, net weight, density
- Include process sequence with multipliers
- Minimum 1 mm machining stock required
- PVF = 3000 by default

#### 8. Bought Parts
For bought parts:
- Use exact cost table entries
- No manual cost overrides allowed
- Match size, grade, length from cost tables

#### 9. Penalty Calculations
- Method A: Direct penalty points
- Method B: Cost penalty = 2 × (Correct Cost - Your Cost)

### Export to PDF
1. Open the workbook in Excel
2. File > Save As > PDF
3. Select all sheets
4. Save as PDF

### Important Notes
- English only
- INR currency (₹)
- 2 decimal places for costs
- Do not modify formula cells (protected)
- Do not reorder sheets
- Follow commodity sequence: BR, EN, FR, EL, MS, ST, SU, WT

### Support
For questions, refer to:
- Appendix_C1_CostMethodology.pdf
- Cost Report Guidelines.pdf
- Cost tables provided
"""
    
    with open('USER_GUIDE.md', 'w') as f:
        f.write(guide_content)
    
    print("User guide created: USER_GUIDE.md")

def main():
    """Main function"""
    excel_file = 'SUPRA_SAEINDIA_Cost_Report_Javitron_2025-26.xlsx'
    pdf_file = 'SUPRA_SAEINDIA_Cost_Report_Javitron_2025-26.pdf'
    
    print("=" * 60)
    print("SUPRA SAEINDIA Cost Report Workbook - Final Steps")
    print("=" * 60)
    
    # Verify workbook structure
    verify_workbook_structure(excel_file)
    
    print()
    
    # Create user guide
    print("Creating user guide...")
    create_user_guide()
    
    print()
    
    # PDF export
    print("PDF Export:")
    print("-" * 60)
    print("To export to PDF manually:")
    print("1. Open the workbook in Excel")
    print("2. File > Save As > PDF")
    print("3. Select all sheets")
    print("4. Save as PDF")
    print()
    print("Note: Automated PDF export requires Microsoft Excel for Mac")
    print("If Excel is installed, the script can attempt automatic export")
    print()
    
    # Try automatic export if Excel is available
    response = input("Attempt automatic PDF export? (y/n): ")
    if response.lower() == 'y':
        print("Attempting automatic PDF export...")
        export_excel_to_pdf_macos(excel_file, pdf_file)
    else:
        print("Skipping automatic export. Please export manually.")
    
    print()
    print("=" * 60)
    print("Workbook is ready for use!")
    print("=" * 60)
    print()
    print("Deliverables:")
    print(f"  ✓ Excel Workbook: {excel_file}")
    print(f"  ✓ User Guide: USER_GUIDE.md")
    print(f"  ✓ PDF: {pdf_file} (export manually or automatically)")
    print()
    print("Next steps:")
    print("  1. Fill in Car Number on Cover Page")
    print("  2. Complete BOM with all parts")
    print("  3. Fill in commodity sheets with parts, materials, processes")
    print("  4. Review validation highlights (red = error, yellow = warning)")
    print("  5. Export to PDF for submission")
    print()

if __name__ == '__main__':
    main()
