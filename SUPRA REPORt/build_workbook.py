#!/usr/bin/env python3
"""
SUPRA SAEINDIA Cost Report Workbook Builder
Creates a cost report workbook matching the Sample FORMATe.xlsx structure
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Team identity
UNIVERSITY_NAME = "Ajeenkya DY Patil University"
TEAM_NAME = "Javitron"
SEASON_YEAR = "2025-26"
VEHICLE_CLASS = "Formula Student / SUPRA"

# Commodity codes and their full names
COMMODITIES = {
    'BR': 'Brakes',
    'EN': 'Engine & Drivetrain',
    'FR': 'Frame & Body',
    'EL': 'Electrical',
    'MS': 'Misc',
    'ST': 'Steering',
    'SU': 'Suspension',
    'WT': 'Wheels'
}

# Assembly mappings per commodity
ASSEMBLY_MAPPING = {
    'BR': ['Brake Pedal', 'Master Cylinder', 'Brake Lines', 'Calipers', 'Rotors'],
    'EN': ['Engine Block', 'Intake', 'Exhaust', 'Gearbox', 'Differential', 'Chain / Sprocket'],
    'FR': ['Main Frame', 'Roll Hoop', 'Impact Structure', 'Belly Panel', 'Firewall', 'Body Panels'],
    'EL': ['Wiring Harness', 'ECU', 'Sensors', 'Battery', 'Switches'],
    'MS': ['Seat', 'Fire Extinguisher', 'Paint', 'Safety Items', 'Fasteners'],
    'ST': ['Steering Wheel', 'Steering Column', 'Rack & Pinion', 'Tie Rods'],
    'SU': ['A-Arms', 'Uprights', 'Push / Pull Rod', 'Dampers', 'Anti-roll Bar'],
    'WT': ['Wheels', 'Tires', 'Bearings', 'Hub']
}

def create_header_style():
    """Create header cell style"""
    return Font(name='Arial', size=11, bold=True, color='FFFFFF')

def create_header_fill():
    """Create header fill style"""
    return PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

def create_border():
    """Create border style"""
    thin = Side(border_style='thin', color='000000')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def create_cover_page(wb):
    """Create the cover page sheet"""
    ws = wb.create_sheet('Cover Page', 0)
    
    # Title
    ws['B2'] = 'SUPRA SAEINDIA Cost Report'
    ws['B2'].font = Font(name='Arial', size=20, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center')
    
    # Team Information
    ws.merge_cells('B5:D5')
    ws['B5'] = 'University:'
    ws['B5'].font = Font(bold=True)
    ws['E5'] = UNIVERSITY_NAME
    
    ws.merge_cells('B7:D7')
    ws['B7'] = 'Team Name:'
    ws['B7'].font = Font(bold=True)
    ws['E7'] = TEAM_NAME
    
    ws.merge_cells('B9:D9')
    ws['B9'] = 'Car Number:'
    ws['B9'].font = Font(bold=True)
    ws['E9'] = ''  # To be filled by user
    
    ws.merge_cells('B11:D11')
    ws['B11'] = 'Season / Year:'
    ws['B11'].font = Font(bold=True)
    ws['E11'] = SEASON_YEAR
    
    ws.merge_cells('B13:D13')
    ws['B13'] = 'Vehicle Class:'
    ws['B13'].font = Font(bold=True)
    ws['E13'] = VEHICLE_CLASS
    
    ws.merge_cells('B15:D15')
    ws['B15'] = 'Report Date:'
    ws['B15'].font = Font(bold=True)
    ws['E15'] = datetime.now().strftime('%Y-%m-%d')
    
    # Column widths
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

def create_toc(wb):
    """Create Table of Contents sheet"""
    ws = wb.create_sheet('Table of Contents', 1)
    
    # Header
    ws['B2'] = 'Table of Contents'
    ws['B2'].font = Font(name='Arial', size=16, bold=True)
    
    # List of sheets
    row = 5
    ws[f'B{row}'] = 'Sheet Name'
    ws[f'C{row}'] = 'Description'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'].font = Font(bold=True)
    
    row += 2
    ws[f'B{row}'] = 'Cover Page'
    ws[f'C{row}'] = 'Team information and report details'
    row += 1
    ws[f'B{row}'] = 'Table of Contents'
    ws[f'C{row}'] = 'Navigation'
    row += 1
    ws[f'B{row}'] = 'Cost Summary'
    ws[f'C{row}'] = 'Overall cost breakdown'
    row += 1
    ws[f'B{row}'] = 'BOM'
    ws[f'C{row}'] = 'Bill of Materials'
    row += 1
    
    for code, name in COMMODITIES.items():
        ws[f'B{row}'] = f'{code} - {name}'
        ws[f'C{row}'] = f'{name} commodity details'
        row += 1
    
    # Column widths
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

def create_cost_summary(wb):
    """Create Cost Summary sheet"""
    ws = wb.create_sheet('Cost Summary', 2)
    
    # Header
    ws['B2'] = 'Cost Summary'
    ws['B2'].font = Font(name='Arial', size=16, bold=True)
    
    # Summary table
    headers = ['Commodity', 'Code', 'Total Cost (₹)']
    row = 5
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = create_header_style()
        cell.fill = create_header_fill()
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for code, name in COMMODITIES.items():
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=code)
        ws.cell(row=row, column=4, value='=0')  # Placeholder formula
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = create_border()
        row += 1
    
    # Total row
    ws.merge_cells(f'B{row}:C{row}')
    ws[f'B{row}'] = 'GRAND TOTAL'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'D{row}'] = '=SUM(D6:D15)'
    ws[f'D{row}'].font = Font(bold=True)
    for col in range(2, 5):
        ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=col).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20

def create_commodity_sheet(wb, commodity_code, commodity_name):
    """Create a commodity sheet based on the sample format"""
    sheet_name = f'{commodity_code} - {commodity_name}'
    ws = wb.create_sheet(sheet_name)
    
    row = 1
    
    # Header section
    ws[f'B{row}'] = 'University:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = UNIVERSITY_NAME
    
    row += 1
    ws[f'B{row}'] = 'Team Name:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = TEAM_NAME
    
    row += 1
    ws[f'B{row}'] = 'System:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = commodity_name
    
    row += 1
    ws[f'B{row}'] = 'Assembly:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = ''  # To be filled
    
    row += 1
    ws[f'B{row}'] = 'P/N Base:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f'{commodity_code}-XX-001'
    
    row += 1
    ws[f'B{row}'] = 'Suffix:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = ''
    
    row += 1
    ws[f'B{row}'] = 'Details:'
    ws[f'B{row}'].font = Font(bold=True)
    ws.merge_cells(f'C{row}:H{row}')
    ws[f'C{row}'] = ''
    
    row += 2
    
    # Cost summary row
    ws[f'B{row}'] = 'Car #:'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = ''
    
    ws[f'D{row}'] = 'Asm Cost:'
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'] = '=0'
    ws[f'E{row}'].number_format = '₹#,##0.00'
    
    ws[f'F{row}'] = 'Qty:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = '1'
    
    ws[f'H{row}'] = 'Extended Cost:'
    ws[f'H{row}'].font = Font(bold=True)
    ws[f'I{row}'] = '=E*G'
    ws[f'I{row}'].number_format = '₹#,##0.00'
    
    row += 2
    
    # Parts section
    ws[f'B{row}'] = 'Parts'
    ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
    
    row += 1
    parts_headers = ['ItemOrder', 'Part', 'Part Cost', 'Quantity', 'Sub Total']
    for col, header in enumerate(parts_headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = create_header_style()
        cell.fill = create_header_fill()
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    # Add 10 empty rows for parts
    for i in range(10):
        row += 1
        ws.cell(row=row, column=2, value=i+1)
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row, column=6).number_format = '₹#,##0.00'
    
    row += 2
    
    # Materials section
    ws[f'B{row}'] = 'Materials'
    ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
    
    row += 1
    material_headers = ['ItemOrder', 'Material', 'Use', 'UnitCost', 'Unit', 'Size1', 'Unit1', 'Size2', 'Unit2', 'Area', 'Length', 'Density', 'Quantity', 'Sub Total']
    for col, header in enumerate(material_headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(size=9, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    # Add 10 empty rows for materials
    for i in range(10):
        row += 1
        ws.cell(row=row, column=2, value=i+1)
        for col in range(2, 16):
            ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row, column=15).number_format = '₹#,##0.00'
    
    row += 2
    
    # Processes section
    ws[f'B{row}'] = 'Processes'
    ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
    
    row += 1
    process_headers = ['ItemOrder', 'Process', 'Use', 'UnitCost', 'Unit', 'Quantity', 'Multiplier', 'Mult. Val.', 'Sub Total']
    for col, header in enumerate(process_headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(size=9, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    # Add 10 empty rows for processes
    for i in range(10):
        row += 1
        ws.cell(row=row, column=2, value=i+1)
        for col in range(2, 11):
            ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row, column=10).number_format = '₹#,##0.00'
    
    row += 2
    
    # Fasteners section
    ws[f'B{row}'] = 'Fasteners'
    ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
    
    row += 1
    fastener_headers = ['ItemOrder', 'Fastener', 'Use', 'UnitCost', 'Size1', 'Unit1', 'Size2', 'Unit2', 'Quantity', 'Sub Total']
    for col, header in enumerate(fastener_headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(size=9, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    # Add 10 empty rows for fasteners
    for i in range(10):
        row += 1
        ws.cell(row=row, column=2, value=i+1)
        for col in range(2, 12):
            ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row, column=11).number_format = '₹#,##0.00'
    
    row += 2
    
    # Tooling section
    ws[f'B{row}'] = 'Tooling'
    ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
    
    row += 1
    tooling_headers = ['ItemOrder', 'Tooling', 'Use', 'UnitCost', 'Unit', 'Quantity', 'PVF', 'FractionIncluded', 'Sub Total']
    for col, header in enumerate(tooling_headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(size=9, bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    # Add 5 empty rows for tooling
    for i in range(5):
        row += 1
        ws.cell(row=row, column=2, value=i+1)
        for col in range(2, 11):
            ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        ws.cell(row=row, column=10).number_format = '₹#,##0.00'
    
    # Set column widths
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 12

def create_bom_sheet(wb):
    """Create BOM sheet"""
    ws = wb.create_sheet('BOM')
    
    # Header
    ws['B2'] = 'Bill of Materials'
    ws['B2'].font = Font(name='Arial', size=16, bold=True)
    
    # BOM headers
    headers = ['Part Number', 'Part Name', 'Assembly', 'Commodity', 'Make/Buy', 'Quantity', 'Unit Cost', 'Extended Cost']
    row = 5
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = create_header_style()
        cell.fill = create_header_fill()
        cell.border = create_border()
        cell.alignment = Alignment(horizontal='center')
    
    # Add 50 empty rows for BOM entries
    for i in range(50):
        row += 1
        for col in range(2, 10):
            ws.cell(row=row, column=col).border = create_border()
        ws.cell(row=row, column=8).number_format = '₹#,##0.00'
        ws.cell(row=row, column=9).number_format = '₹#,##0.00'
    
    # Column widths
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15

def main():
    """Main function to create the workbook"""
    print("Creating SUPRA SAEINDIA Cost Report Workbook...")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create sheets
    print("  - Creating Cover Page...")
    create_cover_page(wb)
    
    print("  - Creating Table of Contents...")
    create_toc(wb)
    
    print("  - Creating Cost Summary...")
    create_cost_summary(wb)
    
    print("  - Creating BOM...")
    create_bom_sheet(wb)
    
    print("  - Creating Commodity Sheets...")
    for code, name in COMMODITIES.items():
        print(f"    - {code} - {name}...")
        create_commodity_sheet(wb, code, name)
    
    # Save workbook
    output_file = 'SUPRA_SAEINDIA_Cost_Report_Javitron_2025-26.xlsx'
    print(f"\nSaving workbook to {output_file}...")
    wb.save(output_file)
    print("Workbook created successfully!")
    
    return output_file

if __name__ == '__main__':
    main()
