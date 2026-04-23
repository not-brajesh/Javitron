#!/usr/bin/env python3
"""
Add advanced validation with conditional formatting to SUPRA SAEINDIA Cost Report Workbook
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.protection import SheetProtection

COMMODITIES = {
    'BR': 'Brakes',
    'EN': 'Engine & Drivetrain',
    'FR': 'Frame & Body',
    'EL': 'Electrical',
    'MS': 'Misc',
    'ST': 'Steering',
    'SU': 'Suspension',
    'WT': 'Wheels'
}

def create_error_fill():
    """Red fill for errors"""
    return PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

def create_warning_fill():
    """Yellow fill for warnings"""
    return PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

def create_border():
    """Create border style"""
    thin = Side(border_style='thin', color='000000')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def add_conditional_formatting_to_commodity(ws):
    """Add conditional formatting to commodity sheets"""
    
    # Red fill for empty required cells
    red_fill = create_error_fill()
    yellow_fill = create_warning_fill()
    
    # Parts section - highlight if Part name is empty (rows 10-20, column C)
    # Formula: =AND($C10<>"", $D10="", $E10="") - yellow if part exists but no cost/qty
    # Formula: =OR($C10="", $D10="", $E10="") - red if any required field is empty
    
    # Add conditional formatting for empty Part names
    ws.conditional_formatting.add(
        'C10:C20',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=red_fill)
    )
    
    # Add conditional formatting for empty costs when part exists
    ws.conditional_formatting.add(
        'D10:D20',
        FormulaRule(formula=['AND(C10<>"", D10="")'], stopIfTrue=True, fill=yellow_fill)
    )
    
    # Materials section - highlight if Material is empty (rows 22-32)
    ws.conditional_formatting.add(
        'C22:C32',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=red_fill)
    )
    
    # Processes section - highlight if Process is empty (rows 34-44)
    ws.conditional_formatting.add(
        'C34:C44',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=red_fill)
    )
    
    # Fasteners section - highlight if Fastener is empty (rows 46-56)
    ws.conditional_formatting.add(
        'C46:C56',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=yellow_fill)
    )
    
    # Tooling section - highlight if Tooling is empty (rows 58-63)
    ws.conditional_formatting.add(
        'C58:C63',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=yellow_fill)
    )

def add_data_validation_dropdowns(ws):
    """Add data validation dropdowns"""
    from openpyxl.worksheet.datavalidation import DataValidation
    
    # Make/Buy dropdown
    dv_make_buy = DataValidation(type="list", formula1='"Make,Buy"', allow_blank=True)
    dv_make_buy.error = 'Please select Make or Buy'
    dv_make_buy.errorTitle = 'Invalid Entry'
    dv_make_buy.prompt = 'Select Make or Buy'
    dv_make_buy.promptTitle = 'Make/Buy Classification'
    
    # Commodity dropdown
    commodities_list = ','.join([f'"{code}"' for code in COMMODITIES.keys()])
    dv_commodity = DataValidation(type="list", formula1=f'{commodities_list}', allow_blank=True)
    dv_commodity.error = 'Please select a valid commodity code'
    dv_commodity.errorTitle = 'Invalid Commodity'
    
    return dv_make_buy, dv_commodity

def apply_advanced_formatting(ws):
    """Apply advanced formatting to sheets"""
    
    # Freeze panes at appropriate rows
    ws.freeze_panes = 'A10'
    
    # Set row heights for headers
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[21].height = 20
    ws.row_dimensions[33].height = 20
    ws.row_dimensions[45].height = 20
    ws.row_dimensions[57].height = 20
    
    # Protect the sheet with formula cells locked
    protection = SheetProtection(
        sheet=True,
        objects=True,
        scenarios=True,
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=False,
        selectLockedCells=True,
        selectUnlockedCells=True,
        sort=False,
        autoFilter=True,
        pivotTables=False
    )
    ws.protection = protection
    
    # Lock formula cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula cell
                cell.protection = Protection(locked=True)

def enhance_bom_validation(ws):
    """Add validation to BOM sheet"""
    red_fill = create_error_fill()
    yellow_fill = create_warning_fill()
    
    # Highlight empty Part Numbers
    ws.conditional_formatting.add(
        'B6:B56',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=red_fill)
    )
    
    # Highlight empty Part Names
    ws.conditional_formatting.add(
        'C6:C56',
        CellIsRule(operator='equal', formula=['""'], stopIfTrue=True, fill=red_fill)
    )
    
    # Highlight if Unit Cost is missing when Part exists
    ws.conditional_formatting.add(
        'H6:H56',
        FormulaRule(formula=['AND(C6<>"", H6="")'], stopIfTrue=True, fill=yellow_fill)
    )
    
    # Freeze panes
    ws.freeze_panes = 'A6'
    
    # Auto filter
    ws.auto_filter.ref = 'B5:I56'

def main():
    """Main function to add advanced validation"""
    print("Adding advanced validation and formatting...")
    
    # Load the workbook
    wb = load_workbook('SUPRA_SAEINDIA_Cost_Report_Javitron_2025-26.xlsx')
    
    # Enhance BOM
    print("  - Enhancing BOM validation...")
    enhance_bom_validation(wb['BOM'])
    
    # Enhance commodity sheets
    print("  - Enhancing Commodity Sheets validation...")
    for code in COMMODITIES.keys():
        sheet_name = f'{code} - {COMMODITIES[code]}'
        print(f"    - {sheet_name}...")
        add_conditional_formatting_to_commodity(wb[sheet_name])
        apply_advanced_formatting(wb[sheet_name])
    
    # Save updated workbook
    output_file = 'SUPRA_SAEINDIA_Cost_Report_Javitron_2025-26.xlsx'
    print(f"\nSaving workbook to {output_file}...")
    wb.save(output_file)
    print("Advanced validation and formatting added successfully!")
    
    return output_file

if __name__ == '__main__':
    main()
